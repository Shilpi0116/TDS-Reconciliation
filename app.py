import streamlit as st
import zipfile
import os
import pandas as pd
import re
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Hardcoded password from your original code
ZIP_PASSWORD = b"05092006"

# Streamlit UI
st.title("26AS Reconciliation Tool")

# File Uploaders
st.header("Upload Files")
zip_file = st.file_uploader("Upload ZIP File", type=["zip"])
tally_file = st.file_uploader("Upload Tally Data File", type=["xlsx"])

# Process ZIP file automatically upon upload
if zip_file:
    with st.spinner("Extracting ZIP and converting to Excel..."):
        extract_to = "extracted"
        os.makedirs(extract_to, exist_ok=True)
        
        with zipfile.ZipFile(zip_file, "r") as zip_ref:
            for file in zip_ref.namelist():
                zip_ref.extract(file, path=extract_to, pwd=ZIP_PASSWORD)
        
        text_file = None
        for file in os.listdir(extract_to):
            if file.endswith(".txt"):
                text_file = os.path.join(extract_to, file)
                break
        
        if text_file:
            with open(text_file, "r", encoding="utf-8") as file:
                lines = file.readlines()
            
            lines = [line.strip() for line in lines if line.strip()]
            original_headers = lines[4].split("^")
            header_row = original_headers
            data_rows = [line.split("^") for line in lines[4:]]
            
            corrected_data = []
            for row in data_rows:
                if len(row) < len(header_row):
                    row += [""] * (len(header_row) - len(row))
                elif len(row) > len(header_row):
                    row = row[:len(header_row)]
                corrected_data.append(row)
            
            df_extracted = pd.DataFrame(corrected_data, columns=header_row)
            st.session_state['df_extracted'] = df_extracted
            st.success("ZIP file extracted and converted successfully!")
        else:
            st.error("No text file found in the extracted folder!")
            st.stop()

# Function to format the Excel sheet
def format_excel_sheet(writer, df, sheet_name):
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    border = Border(left=Side(style="thin"), right=Side(style="thin"), 
                    top=Side(style="thin"), bottom=Side(style="thin"))
    align_center = Alignment(horizontal="center", vertical="center")
    
    # Format headers
    for col_num, value in enumerate(df.columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = value
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = align_center
    
    # Format data rows
    for row_num, row_data in enumerate(df.values, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            # Format numeric columns with thousand separators and 2 decimal places
            if isinstance(value, (int, float)) and not pd.isna(value):
                cell.number_format = '#,##0.00'
    
    # Auto-adjust column widths
    for col_num, column in enumerate(df.columns, 1):
        max_length = max(
            len(str(column)),  # Header length
            max((len(str(row[col_num-1])) for row in df.values if len(str(row[col_num-1])) < 50), default=0)  # Max data length
        )
        worksheet.column_dimensions[get_column_letter(col_num)].width = max_length + 5
    
    # Freeze the top row
    worksheet.freeze_panes = "A2"
    
    # Highlight "Difference" column if not zero (for "Differences" sheet)
    if sheet_name == "Differences":
        diff_col = df.columns.get_loc("Difference") + 1
        for row in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row, column=diff_col)
            if cell.value != 0 and cell.value is not None:
                cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")  # Light red

# Process Reconciliation and Download
if st.button("Process Reconciliation"):
    if tally_file and 'df_extracted' in st.session_state:
        with st.spinner("Processing Reconciliation..."):
            df = pd.read_excel(tally_file, skiprows=13)
            df.rename(columns={df.columns[0]: "Particulars"}, inplace=True)
            
            tan_pattern = r"TDS -([^()]+)\s*\((\w{4}\d{5}\w)"
            df[['Party_Name', 'TAN as per Tally']] = df['Particulars'].str.extract(tan_pattern)
            
            no_tan_df = df[df['TAN as per Tally'].isna()].drop(columns=['TAN as per Tally', 'Party_Name'])
            df_tds = df.dropna(subset=['TAN as per Tally']).copy()
            
            debit_column = "Transactions"
            relevant_columns = ['TAN as per Tally', 'Party_Name', debit_column]
            df_tds = df_tds[relevant_columns]
            df_tds[debit_column] = pd.to_numeric(df_tds[debit_column], errors='coerce')
            
            tds_26as_df = pd.DataFrame(st.session_state['df_extracted'][3:], columns=st.session_state['df_extracted'].columns)
            tds_26as_df.columns = tds_26as_df.columns.str.strip()
            
            tan_column = "TAN of Deductor"
            deductor_column = "Name of Deductor"
            amount_column = "Total Amount Paid / Credited(Rs.)"
            tax_deducted_column = "Total Tax Deducted(Rs.)"
            tds_deposited_column = "Total TDS Deposited(Rs.)"
            
            tds_26as_df = tds_26as_df[[tan_column, deductor_column, amount_column, tax_deducted_column, tds_deposited_column]]
            tds_26as_df = tds_26as_df.dropna(how="all")
            tds_26as_df[tan_column] = tds_26as_df[tan_column].astype(str).str.strip()
            tds_26as_df = tds_26as_df[tds_26as_df[tan_column].str.match(r"^[A-Za-z0-9]{10}$", na=False)]
            
            for col in [amount_column, tax_deducted_column, tds_deposited_column]:
                tds_26as_df[col] = pd.to_numeric(tds_26as_df[col], errors="coerce")
            
            tds_26as_df = tds_26as_df.dropna(subset=[amount_column, tax_deducted_column, tds_deposited_column])
            
            df_26as = tds_26as_df.copy()
            df_reconciled = df_26as.merge(df_tds, left_on=tan_column, right_on="TAN as per Tally", how="outer", indicator=True)
            df_reconciled["Difference"] = df_reconciled["Total Tax Deducted(Rs.)"] - df_reconciled[debit_column]
            df_reconciled.rename(columns={"Transactions": "AMT"}, inplace=True)
            df_reconciled[" "] = ""
            
            fully_matched = df_reconciled[(df_reconciled["_merge"] == "both") & (df_reconciled["Difference"] == 0)].copy()
            only_in_26as = df_reconciled[df_reconciled["_merge"] == "left_only"].copy()
            only_in_tds = df_reconciled[df_reconciled["_merge"] == "right_only"].copy()
            differences = df_reconciled[(df_reconciled["_merge"] == "both") & (df_reconciled["Difference"] != 0)].copy()
            
            df_reconciled.drop(columns=["_merge"], inplace=True)
            fully_matched.drop(columns=["_merge"], inplace=True)
            only_in_26as.drop(columns=["_merge"], inplace=True)
            only_in_tds.drop(columns=["_merge"], inplace=True)
            differences.drop(columns=["_merge"], inplace=True)
            
            for df in [df_reconciled, fully_matched, only_in_26as, only_in_tds, differences]:
                df["Remarks"] = ""
            
            columns_order = [
                "Name of Deductor", "TAN of Deductor", "Total Tax Deducted(Rs.)", " ", 
                "Party_Name", "TAN as per Tally", "AMT", "Difference", "Remarks"
            ]
            
            df_reconciled = df_reconciled[columns_order]
            fully_matched = fully_matched[columns_order]
            only_in_26as = only_in_26as[columns_order]
            only_in_tds = only_in_tds[columns_order]
            differences = differences[columns_order]
            
            def add_total_row(df):
                total_values = {col: df[col].sum() if df[col].dtype in ["int64", "float64"] else "" for col in df.columns}
                total_values["Name of Deductor"] = "TOTAL"
                return pd.concat([df, pd.DataFrame([total_values])], ignore_index=True)
            
            df_reconciled = add_total_row(df_reconciled)
            fully_matched = add_total_row(fully_matched)
            only_in_26as = add_total_row(only_in_26as)
            only_in_tds = add_total_row(only_in_tds)
            differences = add_total_row(differences)
            
            # Create Excel file with formatting
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_reconciled.to_excel(writer, sheet_name="Full_Reconciliation", index=False)
                fully_matched.to_excel(writer, sheet_name="Matched", index=False)
                only_in_26as.to_excel(writer, sheet_name="Only_in_26AS", index=False)
                only_in_tds.to_excel(writer, sheet_name="Only_in_Books", index=False)
                differences.to_excel(writer, sheet_name="Differences", index=False)
                
                # Apply formatting to each sheet
                format_excel_sheet(writer, df_reconciled, "Full_Reconciliation")
                format_excel_sheet(writer, fully_matched, "Matched")
                format_excel_sheet(writer, only_in_26as, "Only_in_26AS")
                format_excel_sheet(writer, only_in_tds, "Only_in_Books")
                format_excel_sheet(writer, differences, "Differences")
            
            output.seek(0)
            st.session_state['reconciliation_file'] = output
            st.success("Reconciliation completed successfully!")
    else:
        st.error("Please upload both ZIP file and Tally file!")

# Show download button only if reconciliation file is available
if 'reconciliation_file' in st.session_state:
    st.download_button(
        label="Download Reconciliation",
        data=st.session_state['reconciliation_file'],
        file_name="TDS_Reconciliation.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="download_reconciliation"
    )